"""Excel レポート生成"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEAD_FILL  = PatternFill(fill_type="solid", fgColor="1B3A5C")
HEAD_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
WARN_FILL  = PatternFill(fill_type="solid", fgColor="FF6B6B")
CAUTION_FILL = PatternFill(fill_type="solid", fgColor="FFD700")
OK_FILL    = PatternFill(fill_type="solid", fgColor="E8F5E9")
BORDER     = Border(*[Side(style="thin")] * 4)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _h(ws, row, col, val, w=12):
    c = ws.cell(row=row, column=col, value=val)
    c.font = HEAD_FONT; c.fill = HEAD_FILL
    c.alignment = CENTER; c.border = BORDER
    ws.column_dimensions[get_column_letter(col)].width = w
    return c

def _c(ws, row, col, val, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", size=9)
    c.alignment = align or CENTER; c.border = BORDER
    if fill: c.fill = fill
    return c


def generate_excel(result: dict, out_path: str):
    wb = openpyxl.Workbook()

    # ── Sheet1: イレギュラー一覧 ──────────────────────────────
    ws1 = wb.active
    ws1.title = "イレギュラー一覧"
    headers = ["個人コード","氏名","店舗","身分","出勤日数","有休日数",
               "欠勤日数","病休日数","有休残日数","実働時間","普通残業","イレギュラー項目"]
    widths  = [12,14,18,14,8,8,8,8,10,10,10,40]
    for i,(h,w) in enumerate(zip(headers, widths), 1):
        _h(ws1, 1, i, h, w)

    for row_i, emp in enumerate(result.get("after_irregular", []), 2):
        flags = emp.get("イレギュラー", [])
        fill  = WARN_FILL if len(flags) >= 2 else (CAUTION_FILL if flags else OK_FILL)
        vals  = [
            emp.get("個人コード",""), emp.get("氏名",""), emp.get("店舗",""), emp.get("身分",""),
            emp.get("出勤日数",""), emp.get("有休日数",""),
            emp.get("欠勤日数",""), emp.get("病休日数",""),
            emp.get("有休残日数",""), emp.get("実働時間",""), emp.get("普通残業",""),
            "、".join(flags) if flags else "問題なし",
        ]
        for ci, v in enumerate(vals, 1):
            _c(ws1, row_i, ci, v, fill=fill, align=(LEFT if ci == 12 else CENTER))

    ws1.row_dimensions[1].height = 22
    ws1.freeze_panes = "A2"

    # ── Sheet2: 修正前後差分 ──────────────────────────────────
    ws2 = wb.create_sheet("修正前後 差分")
    h2 = ["個人コード","氏名","店舗","身分","変更数","変更項目（修正前→修正後）"]
    w2 = [12,14,18,14,8,60]
    for i,(h,w) in enumerate(zip(h2,w2),1):
        _h(ws2, 1, i, h, w)

    for ri, d in enumerate(result.get("diff",[]), 2):
        changed = d.get("変更内容", [])
        detail  = "、".join(f"{c['項目']}:{c['修正前']}→{c['修正後']}" for c in changed)
        fill    = WARN_FILL if d["変更数"] >= 4 else (CAUTION_FILL if d["変更数"] > 0 else OK_FILL)
        for ci, v in enumerate([
            d["個人コード"], d["氏名"], d["店舗"], d["身分"], d["変更数"], detail or "変更なし"
        ], 1):
            _c(ws2, ri, ci, v, fill=fill, align=(LEFT if ci == 6 else CENTER))

    ws2.freeze_panes = "A2"

    # ── Sheet3: 社員マスタ ────────────────────────────────────
    ws3 = wb.create_sheet("社員マスタ")
    mh  = ["個人コード","氏名","店舗","身分","所定日時間","所定週時間","所定週日数"]
    mw  = [12,14,18,14,10,10,10]
    for i,(h,w) in enumerate(zip(mh,mw),1):
        _h(ws3, 1, i, h, w)
    for ri, (code, m) in enumerate(result.get("master",{}).items(), 2):
        for ci, v in enumerate([
            code, m["氏名"], m["店舗"], m["身分"],
            m["所定日時間"], m["所定週時間"], m["所定週日数"]
        ], 1):
            _c(ws3, ri, ci, v)

    ws3.freeze_panes = "A2"

    wb.save(out_path)
